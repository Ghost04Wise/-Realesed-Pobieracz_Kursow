import time
import sys
import re
from bs4 import BeautifulSoup
from requests import get
import tkinter as tk
from openpyxl import load_workbook
from tkinter.filedialog import askopenfilename
import requests

window = tk.Tk()
window.title("Pobieracz Kursów 1.0")
bg_colour = 'lightblue'
window.configure(background=bg_colour)
window.wm_iconbitmap('icon.ico')
logo = tk.PhotoImage(file="logo.gif")


# RYSOWANIE MENU:


def draw_window():
    clear_frame()
    menu = tk.Menu(window)
    window.config(menu=menu)
    menu.add_command(label="    ZMIEŃ LOKALIZACJĘ ARKUSZA   ", command=lambda: change_xlsx_local())
    window_height = 310
    window_width = 840
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x_cordinate = int((screen_width / 2) - (window_width / 2))
    y_cordinate = int((screen_height / 2) - (window_height / 2))
    window.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
    window.resizable(False, False)

    right_side = tk.Frame(window, background=bg_colour)
    right_side.pack(side=tk.RIGHT)
    title = tk.Label(right_side, image=logo)
    title.pack()
    space = tk.Label(right_side, text="\n", background=bg_colour)
    space.pack()
    button_go = tk.Button(right_side, activebackground='gold', text='AKTUALIZUJ ARKUSZ',
                          command=lambda: update_everything(), bg='green', fg="yellow",
                          font='Helvetica 11 bold', width=22)
    button_go.pack()

    left_side = tk.Frame(window)
    left_side.pack(side=tk.LEFT)

    button_usd = tk.Button(left_side, activebackground='gold', text='USD / PLN', command=lambda: change_token_menu(1),
                           bg='silver', fg="brown",
                           font='Helvetica 10 bold', width=25)
    button_usd.pack()

    button_eur = tk.Button(left_side, activebackground='gold', text='EUR / PLN', command=lambda: change_token_menu(2),
                           bg='silver', fg="brown",
                           font='Helvetica 10 bold', width=25)
    button_eur.pack()

    button_gbp = tk.Button(left_side, activebackground='gold', text='GBP / PLN', command=lambda: change_token_menu(3),
                           bg='silver', fg="brown",
                           font='Helvetica 10 bold', width=25)
    button_gbp.pack()

    button_gold = tk.Button(left_side, activebackground='gold', text='GOLD / USD', command=lambda: change_token_menu(4),
                            bg='silver', fg="brown",
                            font='Helvetica 10 bold', width=25)
    button_gold.pack()

    button_silver = tk.Button(left_side, activebackground='gold', text='SILVER / USD', command=lambda: change_token_menu(33),
                            bg='silver', fg="brown",
                            font='Helvetica 10 bold', width=25)
    button_silver.pack()

    button_swda = tk.Button(left_side, activebackground='gold', text='SWDA ETF / GBP',
                            command=lambda: change_token_menu(5), bg='silver', fg="brown",
                            font='Helvetica 10 bold', width=25)
    button_swda.pack()

    button_emim = tk.Button(left_side, activebackground='gold', text='EMIM ETF / GBP',
                            command=lambda: change_token_menu(6), bg='silver', fg="brown",
                            font='Helvetica 10 bold', width=25)
    button_emim.pack()

    button_btc = tk.Button(left_side, activebackground='gold', text=ticker(7), command=lambda: change_token_menu(7),
                           bg='silver', fg="brown",
                           font='Helvetica 10 bold', width=25)
    button_btc.pack()

    button_eth = tk.Button(left_side, activebackground='gold', text=ticker(8), command=lambda: change_token_menu(8),
                           bg='silver', fg="brown",
                           font='Helvetica 10 bold', width=25)
    button_eth.pack()

    button_0 = tk.Button(left_side, activebackground='gold', text=ticker(9), command=lambda: change_token_menu(9),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_0.pack()

    button_1 = tk.Button(left_side, activebackground='gold', text=ticker(10), command=lambda: change_token_menu(10),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_1.pack()

    left2_side = tk.Frame(window)
    left2_side.pack(side=tk.LEFT)

    button_2 = tk.Button(left2_side, activebackground='gold', text=ticker(11), command=lambda: change_token_menu(11),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_2.pack()

    button_3 = tk.Button(left2_side, activebackground='gold', text=ticker(12), command=lambda: change_token_menu(12),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_3.pack()

    button_4 = tk.Button(left2_side, activebackground='gold', text=ticker(13), command=lambda: change_token_menu(13),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_4.pack()

    button_5 = tk.Button(left2_side, activebackground='gold', text=ticker(14), command=lambda: change_token_menu(14),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_5.pack()

    button_6 = tk.Button(left2_side, activebackground='gold', text=ticker(15), command=lambda: change_token_menu(15),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_6.pack()

    button_7 = tk.Button(left2_side, activebackground='gold', text=ticker(16), command=lambda: change_token_menu(16),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_7.pack()

    button_8 = tk.Button(left2_side, activebackground='gold', text=ticker(17), command=lambda: change_token_menu(17),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_8.pack()

    button_9 = tk.Button(left2_side, activebackground='gold', text=ticker(18), command=lambda: change_token_menu(18),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_9.pack()

    button_10 = tk.Button(left2_side, activebackground='gold', text=ticker(19), command=lambda: change_token_menu(19),
                          bg='silver', fg="brown",
                          font='Helvetica 10 bold', width=25)
    button_10.pack()

    button_11 = tk.Button(left2_side, activebackground='gold', text=ticker(20), command=lambda: change_token_menu(20),
                          bg='silver', fg="brown",
                          font='Helvetica 10 bold', width=25)
    button_11.pack()

    button_12 = tk.Button(left2_side, activebackground='gold', text=ticker(21), command=lambda: change_token_menu(21),
                          bg='silver', fg="brown",
                          font='Helvetica 10 bold', width=25)
    button_12.pack()

    left3_side = tk.Frame(window)
    left3_side.pack(side=tk.LEFT)

    button_13 = tk.Button(left3_side, activebackground='gold', text=ticker(22), command=lambda: change_token_menu(22),
                          bg='silver', fg="brown",
                          font='Helvetica 10 bold', width=25)
    button_13.pack()

    button_14 = tk.Button(left3_side, activebackground='gold', text=ticker(23), command=lambda: change_token_menu(23),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_14.pack()

    button_15 = tk.Button(left3_side, activebackground='gold', text=ticker(24), command=lambda: change_token_menu(24),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_15.pack()

    button_16 = tk.Button(left3_side, activebackground='gold', text=ticker(25), command=lambda: change_token_menu(25),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_16.pack()

    button_17 = tk.Button(left3_side, activebackground='gold', text=ticker(26), command=lambda: change_token_menu(26),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_17.pack()

    button_18 = tk.Button(left3_side, activebackground='gold', text=ticker(27), command=lambda: change_token_menu(27),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_18.pack()

    button_19 = tk.Button(left3_side, activebackground='gold', text=ticker(28), command=lambda: change_token_menu(28),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_19.pack()

    button_20 = tk.Button(left3_side, activebackground='gold', text=ticker(29), command=lambda: change_token_menu(29),
                         bg='silver', fg="brown",
                         font='Helvetica 10 bold', width=25)
    button_20.pack()

    button_21 = tk.Button(left3_side, activebackground='gold', text=ticker(30), command=lambda: change_token_menu(30),
                          bg='silver', fg="brown",
                          font='Helvetica 10 bold', width=25)
    button_21.pack()

    button_22 = tk.Button(left3_side, activebackground='gold', text=ticker(31), command=lambda: change_token_menu(31),
                          bg='silver', fg="brown",
                          font='Helvetica 10 bold', width=25)
    button_22.pack()

    button_23 = tk.Button(left3_side, activebackground='gold', text=ticker(32), command=lambda: change_token_menu(32),
                          bg='silver', fg="brown",
                          font='Helvetica 10 bold', width=25)
    button_23.pack()


def change_token_menu(id):
    load_track()
    global ident_t
    global sheet_t
    global cell_t
    clear_frame()
    menu = tk.Menu(window)
    window.config(menu=menu)
    menu.add_command(label="    ZMIEŃ LOKALIZACJĘ ARKUSZA   ", command=lambda: change_xlsx_local())
    window_height = 310
    window_width = 840
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x_cordinate = int((screen_width / 2) - (window_width / 2))
    y_cordinate = int((screen_height / 2) - (window_height / 2))
    window.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
    window.resizable(False, False)

    right_side = tk.Frame(window, background=bg_colour)
    right_side.pack(side=tk.RIGHT)
    title = tk.Label(right_side, image=logo)
    title.pack()
    space = tk.Label(right_side, text="\n", background=bg_colour)
    space.pack()
    button_go = tk.Button(right_side, activebackground='gold', text='MENU GŁÓWNE', command=lambda: draw_window(),
                          bg='green',
                          fg="yellow",
                          font='Helvetica 11 bold', width=22)
    button_go.pack()

    left_side = tk.Frame(window, background=bg_colour)
    left_side.pack(side=tk.TOP)
    if not str(id) in ['1', '2', '3', '4', '5', '6', '33']:
        space2 = tk.Label(left_side, text="\n", background=bg_colour)
        space2.pack()
        name = tk.Label(left_side, text="Podaj API id tokena z Coingecko:", fg="brown", font='Helvetica 11 bold',
                        background=bg_colour)
        name.pack()
        ident_t = tk.Entry(left_side, width=50)
        ident_t.pack()
        ident_t.focus_set()
        ident_t.insert(0, ticker(id))
    if str(id) in ['1', '2', '3', '4', '5', '6', '33']:
        if id == 1:
            space2 = tk.Label(left_side, text="\nUSD \ PLN", fg="red", font='Helvetica 11 bold', background=bg_colour)
            space2.pack()
        if id == 2:
            space2 = tk.Label(left_side, text="\nEUR \ PLN", fg="red", font='Helvetica 11 bold', background=bg_colour)
            space2.pack()
        if id == 3:
            space2 = tk.Label(left_side, text="\nGBP \ PLN", fg="red", font='Helvetica 11 bold', background=bg_colour)
            space2.pack()
        if id == 4:
            space2 = tk.Label(left_side, text="\nGOLD \ USD", fg="red", font='Helvetica 11 bold', background=bg_colour)
            space2.pack()
        if id == 5:
            space2 = tk.Label(left_side, text="\nSWDA ETF \ GBP", fg="red", font='Helvetica 11 bold',
                              background=bg_colour)
            space2.pack()
        if id == 6:
            space2 = tk.Label(left_side, text="\nEMIM ETF \ GBP", fg="red", font='Helvetica 11 bold',
                              background=bg_colour)
            space2.pack()
        if id == 33:
            space2 = tk.Label(left_side, text="\nSILVER \ USD", fg="red", font='Helvetica 11 bold',
                              background=bg_colour)
            space2.pack()

    name1 = tk.Label(left_side, text="Wybierz arkusz w zeszycie:", fg="brown", font='Helvetica 11 bold',
                     background=bg_colour)
    name1.pack()

    sheet_base = wb.sheetnames
    sheet_base.remove('data')
    sheet_t = tk.StringVar(left_side)
    if sheet(id) == 'None' or sheet(id) == '':
        sheet_t.set(sheet_base[0])
    else:
        sheet_t.set(sheet(id))
    sheet1 = tk.OptionMenu(left_side, sheet_t, *sheet_base)
    sheet1.pack()

    name2 = tk.Label(left_side, text="Podaj komórkę w arkuszu (np. A1):", fg="brown", font='Helvetica 11 bold',
                     background=bg_colour)
    name2.pack()
    cell_t = tk.Entry(left_side, width=10)
    cell_t.pack()
    cell_t.insert(0, cell(id))
    space = tk.Label(left_side, text="\n", background=bg_colour)
    space.pack()
    button_get = tk.Button(left_side, activebackground='gold', text='ZMIEŃ / DODAJ', command=lambda: get_token(id),
                           bg='green',
                           fg="yellow",
                           font='Helvetica 11 bold', width=15)
    button_get.pack()
    button_del = tk.Button(left_side, activebackground='gold', text='USUŃ', command=lambda: delete_token(id),
                           bg='green',
                           fg="yellow",
                           font='Helvetica 11 bold', width=15)
    button_del.pack()


def clear_frame():
    for obiekty in window.winfo_children():
        obiekty.destroy()


def ticker(id):
    data = wb['data']
    ticker = data.cell(row=1, column=id)
    if str(ticker.value) == 'None':
        return 'TWÓJ TOKEN'
    return str(ticker.value)


def sheet(id):
    data = wb['data']
    ticker = data.cell(row=2, column=id)
    if str(ticker.value) == 'None':
        return ''
    return str(ticker.value)


def cell(id):
    data = wb['data']
    ticker = data.cell(row=3, column=id)
    if str(ticker.value) == 'None':
        return ''
    return str(ticker.value)


# EDYCJA TOKENÓW:


def get_token(id):
    def ticker(id):
        data = wb['data']
        ticker = data.cell(row=1, column=id)
        if str(ticker.value) == 'None':
            return 'TWÓJ TOKEN'
        return str(ticker.value)
    temp = ticker(id)
    if not str(id) in ['1', '2', '3', '4', '5', '6', '33']:
            ticker = ident_t.get()
            response = requests.get('https://api.coingecko.com/api/v3/simple/price?ids=' + ticker + '&v'
                                                                                                    's_currencies=usd')
            price = response.json()
            check = len(price)
            if check == 1:
                None
            else:
                ticker = temp
    try:
        sheet = sheet_t.get()
        cell = cell_t.get()
        if re.match(r"[a-zA-Z][1-9]$", str(cell)) or re.match(r"[a-zA-Z][1-9][0-9]$", str(cell))\
                or re.match(r"[a-zA-Z][1-9][0-9][0-9]$", str(cell)):
            data = wb['data']
            if not str(id) in ['1', '2', '3', '4', '5', '6', '33']:
                data.cell(row=1, column=id).value = ticker
                data.cell(row=2, column=id).value = sheet
                data.cell(row=3, column=id).value = cell
            else:
                data.cell(row=2, column=id).value = sheet
                data.cell(row=3, column=id).value = cell
            wb.save(trak)
            draw_window()
        else:
            None
    except:
        None


def delete_token(id):
    data = wb['data']
    data.cell(row=1, column=id, value='TWÓJ TOKEN')
    data.cell(row=2, column=id, value='')
    data.cell(row=3, column=id, value='')
    wb.save(trak)
    draw_window()


# ZBIERANIE NOTOWAŃ:


def get_token_price_from_coingecko(id):
    try:
        data = wb['data']
        ticker = data.cell(row=1, column=id)
        ticker = str(ticker.value)

        response = requests.get('https://api.coingecko.com/api/v3/simple/price?ids='+ticker+'&vs_currencies=usd')
        price = response.json()
        price = price[str(ticker)]['usd']
        price = str(price)
        price_exact = price.replace(".", ",")

        data = wb['data']
        sheet = data.cell(row=2, column=id)
        cell = data.cell(row=3, column=id)
        sheet_exact = wb[str(sheet.value)]
        sheet_exact[str(cell.value)] = price_exact
        wb.save(trak)
        print(id)
    except:
        None


def get_fiat_price(link, id):
    try:
        url = link
        page = get(url)
        bs = BeautifulSoup(page.content, 'html.parser')
        c = 0

        for nastronie in bs.find_all('div', class_='left'):
            if c == 0:
                price = nastronie.find('span', itemprop='price')
                price = str(price)
                price = price.replace('<span content="', '')
                price = price.replace('" itemprop="price">', '')
                price = price.replace('</span>', '')
                price = price[0:6]
                price = price.replace('.', ',')
                c += 1

        data = wb['data']
        sheet = data.cell(row=2, column=id)
        cell = data.cell(row=3, column=id)
        sheet_exact = wb[str(sheet.value)]
        sheet_exact[str(cell.value)] = price
        wb.save(trak)
        print(id)
    except:
        None


def get_metal_price(link, id):
    try:
        global notowania
        url = link
        page = get(url)
        bs = BeautifulSoup(page.content, 'html.parser')

        for nastronie in bs.find_all('div', class_='data-blk bid'):
            price = nastronie.find('span').get_text()
            price = price.replace(",", "")
            price = price.replace(".", ",")

        data = wb['data']
        sheet = data.cell(row=2, column=id)
        cell = data.cell(row=3, column=id)
        sheet_exact = wb[str(sheet.value)]
        sheet_exact[str(cell.value)] = price
        wb.save(trak)
        print(id)
    except:
        None


def get_etf_price(link, id):
    try:
        page = get(link)
        bs = BeautifulSoup(page.content, 'html.parser')

        for onpage in bs.find('span', class_='bid price-divide'):
            page = str(onpage)
            page = page.replace('p', '')
            page = page.replace(',', '')
            price = page[0:4]
            print(price)

        data = wb['data']
        sheet = data.cell(row=2, column=id)
        cell = data.cell(row=3, column=id)
        sheet_exact = wb[str(sheet.value)]
        sheet_exact[str(cell.value)] = price
        wb.save(trak)
    except:
        None


def update_everything():
    load_track()
    get_fiat_price('https://e-kursy-walut.pl/kurs-dolara/', 1)
    get_fiat_price('https://e-kursy-walut.pl/kurs-euro/', 2)
    get_fiat_price('https://e-kursy-walut.pl/kurs-funta/', 3)
    get_metal_price('https://www.kitco.com/charts/livegold.html', 4)
    get_metal_price('https://www.kitco.com/charts/livesilver.html', 33)
    get_etf_price('https://www.hl.co.uk/shares/shares-search-results/i/ishares-iii-plc-core-msci-world-acc', 5)
    get_etf_price('https://www.hl.co.uk/shares/shares-search-results/i/ishares-plc-msci-emerging-markets-imi', 6)
    get_token_price_from_coingecko(7)
    get_token_price_from_coingecko(8)
    get_token_price_from_coingecko(9)
    get_token_price_from_coingecko(10)
    get_token_price_from_coingecko(11)
    get_token_price_from_coingecko(12)
    get_token_price_from_coingecko(13)
    get_token_price_from_coingecko(14)
    get_token_price_from_coingecko(15)
    get_token_price_from_coingecko(16)
    get_token_price_from_coingecko(17)
    get_token_price_from_coingecko(18)
    get_token_price_from_coingecko(19)
    get_token_price_from_coingecko(20)
    get_token_price_from_coingecko(21)
    get_token_price_from_coingecko(22)
    get_token_price_from_coingecko(23)
    get_token_price_from_coingecko(24)
    get_token_price_from_coingecko(25)
    get_token_price_from_coingecko(26)
    get_token_price_from_coingecko(27)
    get_token_price_from_coingecko(28)
    get_token_price_from_coingecko(29)
    get_token_price_from_coingecko(30)
    get_token_price_from_coingecko(31)
    get_token_price_from_coingecko(32)
    sys.exit()


# OPERACJE ZE ŚCIEŻKĄ DO ARKUSZA:


def load_track():
    global trak
    try:
        temp = open('data.txt', 'r')
        trak = temp.readline()
        temp.close()
    except:
        check_file_xlsx


def check_file_xlsx():
    try:
        global wb
        wb = load_workbook(trak)
        draw_window()
    except:
        global track
        global ident
        clear_frame()
        window_height = 310
        window_width = 840
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        x_cordinate = int((screen_width / 2) - (window_width / 2))
        y_cordinate = int((screen_height / 2) - (window_height / 2))
        window.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
        window.resizable(False, False)

        right_side = tk.Frame(window, background=bg_colour)
        right_side.pack(side=tk.RIGHT)
        title = tk.Label(right_side, image=logo)
        title.pack()
        space = tk.Label(right_side, text="\n", background=bg_colour)
        space.pack()
        left_side = tk.Frame(window, background=bg_colour)
        left_side.pack(side=tk.LEFT)
        name = tk.Label(left_side, text="    Podaj pełną ścieżkę do arkusza w formacie XLSX:", fg="brown",
                        font='Helvetica 11 bold', background=bg_colour)
        name.pack()
        track1 = askopenfilename(title="Wybierz zeszyt w formacie XLSX")
        track = tk.Entry(left_side, width=50)
        track.pack()
        track.focus_set()
        track.insert(0, track1)
        but = tk.Button(left_side, activebackground='gold', text='...', command=lambda: check_file_xlsx(),
                               bg='green', fg="yellow", font='Helvetica 11 bold', width=4)
        but.pack()
        space = tk.Label(left_side, text="\n", background=bg_colour)
        space.pack()
        button_get = tk.Button(left_side, activebackground='gold', text='DODAJ ŚCIEŻKĘ', command=lambda: add_track(),
                               bg='green',
                               fg="yellow",
                               font='Helvetica 11 bold', width=20)
        button_get.pack()


def change_xlsx_local():
    global track
    global ident
    clear_frame()
    menu = tk.Menu(window)
    window.config(menu=menu)
    menu.add_command(label="    ZMIEŃ LOKALIZACJĘ ARKUSZA   ", command=lambda: change_xlsx_local())
    window_height = 310
    window_width = 840
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x_cordinate = int((screen_width / 2) - (window_width / 2))
    y_cordinate = int((screen_height / 2) - (window_height / 2))
    window.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))
    window.resizable(False, False)

    right_side = tk.Frame(window, background=bg_colour)
    right_side.pack(side=tk.RIGHT)
    title = tk.Label(right_side, image=logo)
    title.pack()
    space = tk.Label(right_side, text="\n", background=bg_colour)
    space.pack()
    button_go = tk.Button(right_side, activebackground='gold', text='MENU GŁÓWNE', command=lambda: draw_window(),
                          bg='green',
                          fg="yellow",
                          font='Helvetica 11 bold', width=22)
    button_go.pack()
    left_side = tk.Frame(window, background=bg_colour)
    left_side.pack(side=tk.LEFT)
    name = tk.Label(left_side, text="    Podaj pełną ścieżkę do arkusza w formacie XLSX:", fg="brown",
                    font='Helvetica 11 bold', background=bg_colour)
    name.pack()
    time.sleep(0.5)
    track1 = askopenfilename(title="Wybierz zeszyt w formacie XLSX")
    track = tk.Entry(left_side, width=50)
    track.pack()
    track.focus_set()
    track.insert(0, track1)
    space = tk.Label(left_side, text="\n", background=bg_colour)
    space.pack()
    button_get = tk.Button(left_side, activebackground='gold', text='DODAJ ŚCIEŻKĘ', command=lambda: add_track(),
                           bg='green',
                           fg="yellow",
                           font='Helvetica 11 bold', width=20)
    button_get.pack()


def add_track():
    global wb
    try:
        tracker = track.get()
        wb = load_workbook(tracker)
        temp = open('data.txt', 'w')
        temp.write(tracker)
        if 'data' in wb.sheetnames:
            None
        else:
            wb.create_sheet('data')
            hidden = wb['data']
            hidden.sheet_state = 'hidden'
            wb.save(tracker)
        wb = load_workbook(tracker)
        draw_window()
    except:
        try:
            load_track()
        except:
            check_file_xlsx()


# WYWOŁANIE PROGRAMU:


load_track()
check_file_xlsx()
window.mainloop()
